from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


HEADER_MAPPING = {
    "import_time": "导入时间",
    "snapshot_date": "快照日期",
    "stock_name": "股票名称",
    "stock_code": "股票代码",
    "avg_buy_price": "买入均价",
    "first_buy_date": "首次买入日期",
    "app_source": "图片来源",
    "current_price": "当前价格",
    "pct_change": "涨跌幅",
    "holding_pnl_amount": "持有盈亏金额",
    "holding_pnl_pct": "持有盈亏比例",
    "daily_pnl_amount": "当日盈亏金额",
    "daily_pnl_pct": "当日盈亏比例",
    "market_value": "持有市值",
    "share_count": "持有股数",
    "holding_days": "持仓天数",
    "position_pct": "个股仓位",
    "realized_pnl": "已实现盈亏",
    "fees": "交易税费",
    "tag": "标签",
    "note": "备注",
    "bs_signal": "BS信号",
    "chart_low": "图区间低点",
    "chart_high": "图区间高点",
    "image_file": "图片文件",
    "parse_confidence": "解析置信度",
    "review_status": "复核状态",
    "latest_buy_date": "最近一次买入日期",
    "total_shares_known": "已知持股数",
    "cost_source": "成本来源",
    "last_updated_at": "最后更新时间",
    "trade_date": "交易日期",
    "trade_type": "交易类型",
    "price": "成交价格",
    "shares": "成交股数",
    "amount": "成交金额",
    "source": "数据来源",
    "signal_type": "信号类型",
    "signal_value": "信号值",
    "field_name": "字段名称",
    "issue_type": "问题类型",
    "issue_detail": "问题详情",
    "severity": "严重程度",
}

SHEET_HEADERS = {
    "holdings_snapshots": [
        "股票名称", "股票代码", "买入均价", "当前价格", "持有盈亏金额", "持有盈亏比例",
        "持有市值", "持有股数", "持仓天数", "个股仓位", "已实现盈亏", "交易税费",
        "BS信号", "图区间低点", "图区间高点",
    ],
    "position_costs": [
        "股票名称", "股票代码", "买入均价", "最近一次买入日期", "已知持股数",
        "成本来源", "最后更新时间",
    ],
    "transaction_ledger": [
        "股票名称", "股票代码", "交易日期", "交易类型", "成交价格", "成交股数", "成交金额",
        "交易税费", "数据来源", "导入时间",
    ],
    "trade_clues": [
        "股票名称", "股票代码", "快照日期", "信号类型", "信号值", "图区间低点", "图区间高点",
        "数据来源", "导入时间",
    ],
    "parse_issues": [
        "股票名称", "股票代码", "字段名称", "问题类型", "问题详情", "严重程度", "数据来源", "导入时间",
    ],
}

FIELD_VALUE_MAPPING = {
    "stock_code": "股票代码",
    "avg_buy_price": "买入均价",
    "first_buy_date": "首次买入日期",
    "snapshot_date": "快照日期",
    "market_value": "持有市值",
}


def normalize_stock_name(raw_name: object) -> str:
    if raw_name is None:
        return ""
    stock_name = str(raw_name).strip()
    if "/" in stock_name or "\\" in stock_name:
        stock_name = Path(stock_name).name
    if "." in stock_name:
        stock_name = stock_name.split(".")[0]
    return stock_name


def stock_name_from_image(image_file: object, fallback_name: object = "") -> str:
    raw_image = "" if image_file is None else str(image_file).strip()
    if raw_image:
        file_name = Path(raw_image).name
        if file_name:
            return normalize_stock_name(file_name.split(".")[0])
    return normalize_stock_name(fallback_name)


def main() -> None:
    workbook_path = Path(__file__).with_name("portfolio-tracker.xlsx")
    workbook = load_workbook(workbook_path)

    snapshot_sheet = workbook["holdings_snapshots"]
    old_headers = [cell.value for cell in snapshot_sheet[1]]
    row_dicts = []
    for row in snapshot_sheet.iter_rows(min_row=2, values_only=True):
        row_dicts.append(dict(zip(old_headers, row)))

    snapshot_rows = []
    for row_dict in row_dicts:
        image_file = row_dict.get("图片文件") or row_dict.get("image_file") or ""
        stock_name = stock_name_from_image(
            image_file,
            row_dict.get("股票名称") or row_dict.get("stock_name") or "",
        )
        snapshot_rows.append([
            stock_name,
            row_dict.get("股票代码") or row_dict.get("stock_code") or "",
            row_dict.get("买入均价") or row_dict.get("avg_buy_price") or "",
            row_dict.get("当前价格") or row_dict.get("current_price") or "",
            row_dict.get("持有盈亏金额") or row_dict.get("holding_pnl_amount") or "",
            row_dict.get("持有盈亏比例") or row_dict.get("holding_pnl_pct") or "",
            row_dict.get("持有市值") or row_dict.get("market_value") or "",
            row_dict.get("持有股数") or row_dict.get("share_count") or "",
            row_dict.get("持仓天数") or row_dict.get("holding_days") or "",
            row_dict.get("个股仓位") or row_dict.get("position_pct") or "",
            row_dict.get("已实现盈亏") or row_dict.get("realized_pnl") or "",
            row_dict.get("交易税费") or row_dict.get("fees") or "",
            row_dict.get("BS信号") or row_dict.get("bs_signal") or "",
            row_dict.get("图区间低点") or row_dict.get("chart_low") or "",
            row_dict.get("图区间高点") or row_dict.get("chart_high") or "",
        ])

    position_rows = []
    position_sheet = workbook["position_costs"]
    old_headers = [cell.value for cell in position_sheet[1]]
    row_dicts = []
    for row in position_sheet.iter_rows(min_row=2, values_only=True):
        row_dicts.append(dict(zip(old_headers, row)))
    for row_dict in row_dicts:
        stock_name = stock_name_from_image(
            row_dict.get("图片文件") or row_dict.get("image_file") or "",
            row_dict.get("股票名称") or row_dict.get("stock_name") or "",
        )
        position_rows.append([
            stock_name,
            row_dict.get("股票代码") or row_dict.get("stock_code") or "",
            row_dict.get("买入均价") or row_dict.get("avg_buy_price") or "",
            row_dict.get("最近一次买入日期") or row_dict.get("latest_buy_date") or "",
            row_dict.get("已知持股数") or row_dict.get("total_shares_known") or "",
            row_dict.get("成本来源") or row_dict.get("cost_source") or "",
            row_dict.get("最后更新时间") or row_dict.get("last_updated_at") or "",
        ])

    other_sheet_rows = {}
    for sheet_name in ["transaction_ledger", "trade_clues", "parse_issues"]:
        sheet = workbook[sheet_name]
        rows = []
        for row in sheet.iter_rows(min_row=2, max_col=len(SHEET_HEADERS[sheet_name]), values_only=True):
            rows.append(list(row))
        other_sheet_rows[sheet_name] = rows

    rebuilt = Workbook()
    default_sheet = rebuilt.active
    rebuilt.remove(default_sheet)

    ordered_sheets = ["holdings_snapshots", "position_costs", "transaction_ledger", "trade_clues", "parse_issues"]
    payload = {
        "holdings_snapshots": snapshot_rows,
        "position_costs": position_rows,
        **other_sheet_rows,
    }

    for sheet_name in ordered_sheets:
        sheet = rebuilt.create_sheet(sheet_name)
        headers = SHEET_HEADERS[sheet_name]
        sheet.append(headers)
        for row in payload.get(sheet_name, []):
            normalized = list(row[: len(headers)])
            if len(normalized) < len(headers):
                normalized.extend([""] * (len(headers) - len(normalized)))
            if sheet_name == "parse_issues":
                field_name = normalized[2]
                if field_name in FIELD_VALUE_MAPPING:
                    normalized[2] = FIELD_VALUE_MAPPING[field_name]
            sheet.append(normalized)

        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 36)

    temp_path = workbook_path.with_name("portfolio-tracker.regenerated.xlsx")
    rebuilt.save(temp_path)

    output_path = workbook_path
    try:
        temp_path.replace(workbook_path)
    except PermissionError:
        output_path = temp_path

    verification_workbook = load_workbook(output_path)
    for sheet in verification_workbook.worksheets:
        headers = [cell.value for cell in sheet[1]]
        print(sheet.title)
        print(headers)

    print(f"saved_to={output_path}")


if __name__ == "__main__":
    main()