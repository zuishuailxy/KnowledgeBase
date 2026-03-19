from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import sys
from typing import Iterable

from openpyxl import Workbook, load_workbook
from rapidocr_onnxruntime import RapidOCR

ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from stocks.StockHoldings.update_headers_to_cn import SHEET_HEADERS


WORKBOOK_PATH = Path(__file__).with_name("portfolio-tracker.xlsx")
IMAGE_DIR = Path(__file__).resolve().parents[1] / "imgs"
IMPORT_TIME = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
AMOUNT_LABELS = {"持有盈亏", "当日盈亏", "持有市值", "实现盈亏", "交易税费"}
PERCENT_LABELS = {"持有盈亏比例", "当日盈亏比例", "个股仓位"}
STOCK_CODE_MAP = {
    "贵州茅台": "600519",
    "比亚迪": "002594",
    "中国移动": "600941",
    "中国海油": "600938",
    "五粮液": "000858",
    "亨通光电": "600487",
    "分众传媒": "002027",
    "南山铝业": "600219",
    "德业股份": "605117",
    "恩华药业": "002262",
    "招商银行": "600036",
    "标普500ETF": "513500",
    "格力电器": "000651",
    "海大集团": "002311",
    "海康威视": "002415",
    "海螺水泥": "600585",
    "珀莱雅": "603605",
    "福耀玻璃": "600660",
    "紫金矿业": "601899",
    "美的集团": "000333",
    "顺丰控股": "002352",
}


@dataclass
class SnapshotRecord:
    stock_name: str
    stock_code: str
    avg_buy_price: object
    current_price: object
    holding_pnl_amount: object
    holding_pnl_pct: object
    market_value: object
    share_count: object
    holding_days: object
    position_pct: object
    realized_pnl: object
    fees: object
    bs_signal: str
    chart_low: object
    chart_high: object

    def as_row(self) -> list[object]:
        return [
            self.stock_name,
            self.stock_code,
            self.avg_buy_price,
            self.current_price,
            self.holding_pnl_amount,
            self.holding_pnl_pct,
            self.market_value,
            self.share_count,
            self.holding_days,
            self.position_pct,
            self.realized_pnl,
            self.fees,
            self.bs_signal,
            self.chart_low,
            self.chart_high,
        ]


def normalize_text(text: str) -> str:
    return text.replace(" ", "").replace("|", "").strip()


def file_stem(path: Path) -> str:
    return path.name.split(".")[0]


def lookup_stock_code(stock_name: str) -> str:
    return STOCK_CODE_MAP.get(stock_name, "")


def parse_number(token: str) -> float | int | None:
    cleaned = token.replace(",", "").replace("+", "").replace("#", "")
    cleaned = cleaned.replace("。", ".").replace("..", ".")
    if not cleaned:
        return None
    if cleaned.endswith("万"):
        base = parse_number(cleaned[:-1])
        if base is None:
            return None
        return round(float(base) * 10000, 2)
    if cleaned.endswith("天"):
        cleaned = cleaned[:-1]
    cleaned = cleaned.strip("-")
    if not cleaned:
        return None
    if cleaned.count(".") > 1:
        parts = cleaned.split(".")
        cleaned = "".join(parts[:-1]) + "." + parts[-1]
    try:
        value = float(cleaned)
    except ValueError:
        return None
    if value.is_integer():
        return int(value)
    return round(value, 4)


def parse_percent(token: str) -> float | None:
    cleaned = token.replace("%", "").replace(",", "").replace("+", "")
    cleaned = cleaned.replace("。", ".")
    if cleaned.count(".") > 1:
        head, tail = cleaned.split(".", 1)
        cleaned = head + "." + tail.replace(".", "")
    try:
        return round(float(cleaned) / 100, 4)
    except ValueError:
        return None


def next_numeric_tokens(tokens: list[str], start_index: int, count: int) -> list[str]:
    values: list[str] = []
    for token in tokens[start_index:]:
        if token in AMOUNT_LABELS or token in PERCENT_LABELS:
            continue
        if any(symbol in token for symbol in ("%", "天")):
            continue
        if parse_number(token) is not None:
            values.append(token)
        if len(values) == count:
            break
    return values


def next_percent_tokens(tokens: list[str], start_index: int, count: int) -> list[str]:
    values: list[str] = []
    for token in tokens[start_index:]:
        if token.endswith("%") and parse_percent(token) is not None:
            values.append(token)
        if len(values) == count:
            break
    return values


def find_token(tokens: list[str], keyword: str) -> int | None:
    for index, token in enumerate(tokens):
        if keyword in token:
            return index
    return None


def parse_current_price(token: str) -> tuple[float | int | None, float | None]:
    compact = normalize_text(token).replace("现价", "")
    split_index = max(compact.rfind("+"), compact.rfind("-"))
    if split_index <= 0:
        return parse_number(compact), None
    price_token = compact[:split_index]
    pct_token = compact[split_index:]
    return parse_number(price_token), parse_percent(pct_token)


def parse_chart_range(tokens: list[str], current_price: object) -> tuple[object, object, list[tuple[str, str, str]]]:
    issues: list[tuple[str, str, str]] = []
    start_index = find_token(tokens, "看行情")
    if start_index is None:
        issues.append(("图区间低点", "缺失", "截图中未稳定识别到图区间低点"))
        issues.append(("图区间高点", "缺失", "截图中未稳定识别到图区间高点"))
        return "", "", issues

    candidates: list[float] = []
    for token in tokens[start_index + 1 :]:
        if "/" in token:
            break
        value = parse_number(token)
        if value is None:
            continue
        if isinstance(value, (int, float)) and value > 0:
            candidates.append(float(value))

    if isinstance(current_price, (int, float)) and current_price > 0:
        lower_bound = float(current_price) * 0.1
        upper_bound = float(current_price) * 3
        candidates = [value for value in candidates if lower_bound <= value <= upper_bound]

    if len(candidates) < 2:
        issues.append(("图区间低点", "缺失", "截图中未稳定识别到图区间低点"))
        issues.append(("图区间高点", "缺失", "截图中未稳定识别到图区间高点"))
        return "", "", issues

    low = round(min(candidates), 2)
    high = round(max(candidates), 2)
    issues.append(("图区间低点", "低置信度", f"OCR 推断图区间低点为 {low}"))
    issues.append(("图区间高点", "低置信度", f"OCR 推断图区间高点为 {high}"))
    return low, high, issues


def parse_snapshot(path: Path, engine: RapidOCR) -> tuple[SnapshotRecord, list[tuple[str, str, str]]]:
    result, _ = engine(str(path))
    tokens = [] if not result else [normalize_text(item[1]) for item in result if normalize_text(item[1])]
    stock_name = file_stem(path)
    stock_code = lookup_stock_code(stock_name)

    current_price = ""
    holding_pnl_amount = ""
    holding_pnl_pct = ""
    market_value = ""
    share_count = ""
    holding_days = ""
    position_pct = ""
    realized_pnl = ""
    fees = ""
    bs_signal = ""
    chart_low = ""
    chart_high = ""
    issues: list[tuple[str, str, str]] = [
        ("买入均价", "缺失", "截图未显示买入均价"),
        ("BS信号", "待确认", "截图仅显示查看BS点，未识别到明确 B/S 标记"),
    ]
    if not stock_code:
        issues.insert(0, ("股票代码", "缺失", "截图未显示股票代码，且未命中内置代码映射"))

    for token in tokens:
        if token.startswith("现价"):
            current_price, current_pct = parse_current_price(token)
            if current_pct is not None:
                pass
            break

    pnl_index = find_token(tokens, "持有盈亏")
    if pnl_index is not None:
        amount_tokens = next_numeric_tokens(tokens, pnl_index + 1, 2)
        percent_tokens = next_percent_tokens(tokens, pnl_index + 1, 2)
        if amount_tokens:
            holding_pnl_amount = parse_number(amount_tokens[0]) or ""
        if percent_tokens:
            holding_pnl_pct = parse_percent(percent_tokens[0]) or ""

    market_index = find_token(tokens, "持有市值")
    realized_index = find_token(tokens, "实现盈亏")
    if market_index is not None or realized_index is not None:
        start_index = min(index for index in [market_index, realized_index] if index is not None) + 1
        amount_tokens = next_numeric_tokens(tokens, start_index, 2)
        numeric_values = [parse_number(token) for token in amount_tokens]
        numeric_values = [value for value in numeric_values if value is not None]
        if len(numeric_values) == 2:
            market_value = max(numeric_values)
            realized_pnl = min(numeric_values)
        elif len(numeric_values) == 1:
            market_value = numeric_values[0]

    share_index = find_token(tokens, "持有股数")
    if share_index is not None:
        share_tokens = next_numeric_tokens(tokens, share_index + 1, 1)
        if share_tokens:
            share_count = parse_number(share_tokens[0]) or ""

    position_index = find_token(tokens, "个股仓位")
    if position_index is not None:
        position_tokens = next_percent_tokens(tokens, position_index + 1, 1)
        if position_tokens:
            position_pct = parse_percent(position_tokens[0]) or ""

    holding_days_index = find_token(tokens, "持仓天数")
    if holding_days_index is not None:
        for token in tokens[holding_days_index + 1 :]:
            if token.endswith("天"):
                holding_days = parse_number(token) or ""
                break

    fees_index = find_token(tokens, "交易税费")
    if fees_index is not None:
        fee_tokens = next_numeric_tokens(tokens, fees_index + 1, 1)
        if fee_tokens:
            fees = parse_number(fee_tokens[0]) or ""

    chart_low, chart_high, chart_issues = parse_chart_range(tokens, current_price)
    issues.extend(chart_issues)

    snapshot = SnapshotRecord(
        stock_name=stock_name,
        stock_code=stock_code,
        avg_buy_price="",
        current_price=current_price,
        holding_pnl_amount=holding_pnl_amount,
        holding_pnl_pct=holding_pnl_pct,
        market_value=market_value,
        share_count=share_count,
        holding_days=holding_days,
        position_pct=position_pct,
        realized_pnl=realized_pnl,
        fees=fees,
        bs_signal=bs_signal,
        chart_low=chart_low,
        chart_high=chart_high,
    )
    return snapshot, issues


def ensure_workbook() -> None:
    if WORKBOOK_PATH.exists():
        return

    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, headers in SHEET_HEADERS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
    workbook.save(WORKBOOK_PATH)


def read_existing_rows(sheet) -> set[tuple[object, ...]]:
    rows: set[tuple[object, ...]] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.add(tuple(row))
    return rows


def snapshot_row_key(row: list[object] | tuple[object, ...]) -> tuple[object, ...]:
    current_price = row[3]
    position_pct = row[9]
    if isinstance(current_price, (int, float)):
        current_price = round(float(current_price), 2)
    if isinstance(position_pct, (int, float)):
        position_pct = round(float(position_pct), 4)
    return (row[0], current_price, row[7], row[8], position_pct)


def append_unique_rows(sheet, rows: Iterable[list[object]], key_fn=None) -> int:
    existing_rows = read_existing_rows(sheet)
    existing_keys = {key_fn(list(row)) for row in existing_rows} if key_fn else set()
    inserted = 0
    for row in rows:
        row_tuple = tuple(row)
        row_key = key_fn(row) if key_fn else None
        if row_tuple in existing_rows:
            continue
        if key_fn and row_key in existing_keys:
            continue
        sheet.append(row)
        existing_rows.add(row_tuple)
        if key_fn:
            existing_keys.add(row_key)
        inserted += 1
    return inserted


def build_issue_rows(stock_name: str, stock_code: str, image_name: str, issues: list[tuple[str, str, str]]) -> list[list[object]]:
    return [
        [stock_name, stock_code, field_name, issue_type, issue_detail, "中", image_name, IMPORT_TIME]
        for field_name, issue_type, issue_detail in issues
    ]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--rebuild", action="store_true", help="Recreate workbook from current images before importing")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.rebuild and WORKBOOK_PATH.exists():
        WORKBOOK_PATH.unlink()
    ensure_workbook()
    workbook = load_workbook(WORKBOOK_PATH)
    snapshot_sheet = workbook["holdings_snapshots"]
    trade_clues_sheet = workbook["trade_clues"]
    parse_issues_sheet = workbook["parse_issues"]

    engine = RapidOCR()
    snapshot_rows: list[list[object]] = []
    issue_rows: list[list[object]] = []
    processed_images = 0

    for image_path in sorted(IMAGE_DIR.glob("*.jpg")):
        processed_images += 1
        snapshot, issues = parse_snapshot(image_path, engine)
        snapshot_rows.append(snapshot.as_row())
        issue_rows.extend(build_issue_rows(snapshot.stock_name, snapshot.stock_code, image_path.name, issues))

    inserted_snapshots = append_unique_rows(snapshot_sheet, snapshot_rows, key_fn=snapshot_row_key)
    inserted_trade_clues = append_unique_rows(trade_clues_sheet, [])
    inserted_issues = append_unique_rows(parse_issues_sheet, issue_rows)
    workbook.save(WORKBOOK_PATH)

    print(f"processed_images={processed_images}")
    print(f"inserted_snapshots={inserted_snapshots}")
    print(f"inserted_trade_clues={inserted_trade_clues}")
    print(f"inserted_issues={inserted_issues}")
    print(f"saved_to={WORKBOOK_PATH}")


if __name__ == "__main__":
    main()